"""Reads the evaluation workbook's layout and writes run results back into it.

The cell map is derived from the file itself rather than hardcoded, so a new
release of the workbook that shifts rows around still works: we locate the
header row by its 'Result code' column and the answer blocks by their
'Step N' labels in column A.
"""
from __future__ import annotations

import re
from dataclasses import dataclass

import openpyxl
from openpyxl.utils import get_column_letter

from .client import CallRecord, minify

STEP_RE = re.compile(r"^Step\s*\d+[a-z]?$", re.IGNORECASE)

# Header text in the workbook -> the CallRecord attribute that fills it.
# Matching is substring-based and case-insensitive because the sheets spell
# these inconsistently ('Tag Number  ', 'location Name Created  ', etc).
COLUMN_ALIASES = [
    ("result code", "status"),
    ("license facility", "license_number"),
    ("id number", "object_ids"),
    ("last modified", "last_modified"),
    # Checked before the plain "tag number" alias: the PlantBatches tabs label
    # one column "Plant Batch Name/Tag Number", which takes either.
    ("plant batch name", "tags_or_names"),
    ("tag number", "tags"),
    ("harvest name", "harvest_name"),
    ("name created", "names"),
    ("request sent", "url"),
    ("json body", "evidence"),
]


@dataclass
class SheetMap:
    name: str
    header_row: int
    columns: dict          # semantic field -> column letter
    steps: dict            # 'Step 1' -> row number
    step_tasks: dict       # 'Step 1' -> the task text from column A


def _semantic(header: str) -> str | None:
    low = header.strip().lower()
    for needle, field in COLUMN_ALIASES:
        if needle in low:
            return field
    return None


def map_sheet(ws) -> SheetMap | None:
    header_row = None
    for r in range(1, min(12, ws.max_row + 1)):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(r, c).value or "").strip().lower().startswith("result code"):
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        return None

    columns: dict = {}
    # Column A always carries the task text (and sometimes a state note that
    # would otherwise match a data header), so answers never go there.
    for c in range(2, ws.max_column + 1):
        header = str(ws.cell(header_row, c).value or "")
        field = _semantic(header)
        if field and field not in columns:
            columns[field] = get_column_letter(c)

    steps: dict = {}
    tasks: dict = {}
    for r in range(header_row + 1, ws.max_row + 1):
        label = str(ws.cell(r, 1).value or "").strip()
        if not STEP_RE.match(label):
            continue
        steps[label] = r
        text = []
        for rr in range(r + 1, min(r + 4, ws.max_row + 1)):
            nxt = str(ws.cell(rr, 1).value or "").strip()
            if not nxt or STEP_RE.match(nxt):
                break
            text.append(nxt)
        tasks[label] = " ".join(text)

    return SheetMap(ws.title, header_row, columns, steps, tasks)


def map_workbook(path: str) -> dict:
    wb = openpyxl.load_workbook(path)
    out = {}
    for ws in wb.worksheets:
        sm = map_sheet(ws)
        if sm and sm.steps:
            out[ws.title] = sm
    return out


def _render(field: str, record: CallRecord) -> str:
    if field == "status":
        return str(record.status or "")
    if field == "license_number":
        return record.license_number
    if field == "object_ids":
        return ", ".join(str(i) for i in record.object_ids)
    if field == "tags":
        return ", ".join(str(t) for t in record.tags)
    if field == "tags_or_names":
        # One column serving both: a tag identifies the batch where the state
        # issues one, otherwise its name does.
        return ", ".join(str(v) for v in (record.tags or record.names))
    if field == "names":
        return ", ".join(str(n) for n in record.names)
    if field == "harvest_name":
        return ", ".join(str(n) for n in record.names)
    if field == "last_modified":
        return record.last_modified
    if field == "url":
        return record.url
    if field == "evidence":
        return record.evidence()
    return ""


def write_results(
    src_path: str,
    dest_path: str,
    recorder,
    *,
    company: dict | None = None,
    only_sheets: set | None = None,
) -> dict:
    """Fill the workbook from a run. Returns a per-sheet write count."""
    wb = openpyxl.load_workbook(src_path)
    written: dict = {}

    for ws in wb.worksheets:
        if only_sheets is not None and ws.title not in only_sheets:
            continue
        sm = map_sheet(ws)
        if not sm or not sm.steps:
            continue
        count = 0
        for step, row in sm.steps.items():
            record = recorder.for_step(ws.title, step)
            if record is None:
                continue
            for field, col in sm.columns.items():
                value = _render(field, record)
                if value == "":
                    continue
                ws[f"{col}{row}"] = value
                # Long JSON needs to wrap rather than bleed across the sheet.
                if field == "evidence":
                    ws[f"{col}{row}"].alignment = openpyxl.styles.Alignment(
                        wrap_text=True, vertical="top"
                    )
            count += 1
        if count:
            written[ws.title] = count

    if company:
        _fill_company(wb, company)

    wb.save(dest_path)
    return written


def _fill_company(wb, company: dict) -> None:
    """Fill the CompanyInformation tab by matching its row labels."""
    ws = wb["CompanyInformation"]
    for r in range(1, ws.max_row + 1):
        label = str(ws.cell(r, 2).value or "").strip().lower().rstrip("*").strip()
        if not label:
            continue
        for key, value in company.items():
            if key.strip().lower() == label:
                ws.cell(r, 3).value = value
                break
